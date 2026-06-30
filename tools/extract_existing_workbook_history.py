#!/usr/bin/env python3
"""Extract historical trading rows from the existing paper-trading workbook."""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def clean(value):
    if value is None:
        return ""
    return value


def row_dict(headers, values):
    return {str(header): clean(value) for header, value in zip(headers, values) if header}


def extract(path: Path) -> dict:
    wb = load_workbook(path, data_only=False)
    payload = {"workbook": str(path), "sheets": wb.sheetnames, "plan_rows": [], "execution_rows": [], "summary_rows": []}

    if "股票池与交易计划" in wb.sheetnames:
        ws = wb["股票池与交易计划"]
        headers = [cell.value for cell in ws[7]]
        for row in ws.iter_rows(min_row=8, values_only=True):
            item = row_dict(headers, row)
            symbol = str(item.get("股票代码", "") or "").strip()
            if not symbol or symbol == "现金":
                continue
            payload["plan_rows"].append(item)

    if "盘中执行记录" in wb.sheetnames:
        ws = wb["盘中执行记录"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = row_dict(headers, row)
            if any(value not in ("", None) for value in item.values()):
                payload["execution_rows"].append(item)

    if "收益与准确度汇总" in wb.sheetnames:
        ws = wb["收益与准确度汇总"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = row_dict(headers, row)
            if item.get("指标"):
                payload["summary_rows"].append(item)

    return payload


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: extract_existing_workbook_history.py <workbook.xlsx>", file=sys.stderr)
        return 2
    print(json.dumps(extract(Path(sys.argv[1])), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
