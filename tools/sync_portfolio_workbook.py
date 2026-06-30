#!/usr/bin/env python3
"""Sync the single portfolio ledger CSV into a readable one-sheet workbook."""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


DISPLAY_HEADERS = {
    "date": "日期",
    "section": "分段",
    "row_type": "记录类型",
    "time": "时间",
    "symbol": "股票代码",
    "name": "名称",
    "side": "方向",
    "action": "动作",
    "quantity": "数量",
    "price": "价格",
    "avg_cost": "平均成本",
    "gross_amount": "成交/计划金额",
    "commission": "佣金",
    "stamp_tax": "印花税",
    "transfer_fee": "过户/经手费",
    "total_cost": "交易成本",
    "net_amount": "现金净流",
    "cash": "组合现金",
    "stock_value": "组合市值",
    "market_value": "个股市值",
    "start_equity": "期初权益",
    "end_equity": "期末权益/总资产",
    "daily_pnl": "当日盈亏",
    "daily_return_pct": "当日收益率%",
    "realized_pnl": "已实现盈亏",
    "unrealized_pnl": "未实现盈亏",
    "unrealized_pnl_pct": "未实现收益率%",
    "total_exposure_pct": "总暴露%",
    "weight_pct": "仓位%",
    "max_drawdown_pct": "最大回撤%",
    "risk_state": "风险状态",
    "strategy_tag": "策略标签",
    "strategy_id": "策略ID",
    "score_profile": "评分因子",
    "factor_weights": "因子/策略权重",
    "strategy_status": "策略状态",
    "reason": "买卖/复查理由",
    "thesis": "验证口径",
    "review_date": "后续验证",
    "status": "判断状态",
    "notes": "备注",
}

TYPE_LABELS = {
    "TRADE": "成交",
    "OPENING_POSITION": "期初持仓",
    "TARGET_PORTFOLIO": "目标组合",
    "TRADE_PLAN": "交易计划",
    "DAY_SUMMARY": "组合汇总",
    "POSITION": "持仓",
    "INTRADAY_CHECK": "盘中复查",
    "STRATEGY_FEEDBACK": "策略反馈",
    "BACKTEST": "回测",
    "AUDIT_FINDING": "审计发现",
}

SIDE_LABELS = {
    "BUY": "买入",
    "SELL": "卖出",
    "LONG": "持有",
    "WATCH": "观察",
    "CASH": "现金",
    "RESEARCH": "研究",
}

ACTION_LABELS = {
    "BUY": "买入",
    "SELL": "卖出",
    "HOLD": "持有",
    "WATCH": "观察",
}

STATUS_LABELS = {
    "filled": "已成交",
    "open": "持仓中",
    "completed": "已完成",
    "pending": "待处理",
    "done": "已完成",
    "failed": "失败",
    "waived": "已豁免",
    "overdue": "逾期未留证",
    "待验证": "逾期未留证",
}

TYPE_COLORS = {
    "成交": "E2F0D9",
    "期初持仓": "D9EAD3",
    "目标组合": "C6E0B4",
    "交易计划": "DDEBF7",
    "组合汇总": "FFF2CC",
    "持仓": "FCE4D6",
    "盘中复查": "EADCF8",
    "策略反馈": "D9EAD3",
    "回测": "D9EAF7",
    "审计发现": "F4CCCC",
}

CHECKPOINT_HEADERS = {
    "date": "日期",
    "checkpoint": "检查点",
    "due_time": "截止时间",
    "status": "状态",
    "completed_at": "完成/补登记时间",
    "evidence": "证据",
    "notes": "说明",
}

STATUS_COLORS = {
    "done": "E2F0D9",
    "pending": "DDEBF7",
    "overdue": "FCE4D6",
    "failed": "F4CCCC",
    "waived": "FFF2CC",
}

CONFIG_ROWS = [
    ("INITIAL_CAPITAL", 1000000, "初始训练资金"),
    ("MAX_DRAWDOWN_STOP", 0.05, "触发 STOP 的当前回撤阈值"),
    ("MAX_DRAWDOWN_WARNING", 0.03, "触发 WARNING 的当前回撤阈值"),
    ("DEFAULT_COMMISSION_RATE", 0.0003, "默认佣金率"),
    ("DEFAULT_STAMP_TAX_RATE", 0.0005, "当前账本卖出印花税率"),
    ("DEFAULT_TRANSFER_RATE", 0.00001, "过户/经手费率占位"),
    ("MIN_TRADE_LOT", 100, "A股最小交易单位"),
    ("ODD_LOT_ALLOWED", False, "是否允许卖出制造零股"),
]

ORDER_HEADERS = [
    "order_id",
    "trade_date",
    "symbol",
    "name",
    "side",
    "order_type",
    "quantity",
    "limit_price",
    "exec_price",
    "exec_time",
    "status",
    "strategy_id",
    "signal_score",
    "reason",
    "plan_id",
    "created_at",
]

POSITION_HEADERS = [
    "date",
    "symbol",
    "name",
    "quantity",
    "avg_cost",
    "last_price",
    "market_value",
    "unrealized_pnl",
    "unrealized_pnl_%",
    "weight_%",
    "cost_basis",
    "note",
]

SUMMARY_HEADERS = [
    "date",
    "cash",
    "total_market_value",
    "total_equity",
    "daily_pnl",
    "daily_return_%",
    "cumulative_return_%",
    "max_drawdown_%",
    "current_drawdown_%",
    "var_95%",
    "exposure_%",
    "risk_status",
    "leverage",
    "note",
]

SIGNAL_HEADERS = [
    "signal_id",
    "signal_time",
    "symbol",
    "strategy_id",
    "signal_type",
    "score",
    "entry_threshold",
    "exit_threshold",
    "suggested_action",
    "actual_action",
    "valid",
    "reason",
]

AUDIT_HEADERS = [
    "audit_id",
    "audit_date",
    "issue_type",
    "description",
    "affected_order_id",
    "affected_date",
    "status",
    "fix_action",
    "fix_date",
    "verified_by",
    "raw_detail",
]


def display_rows(rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    headers = rows[0]
    output = [[DISPLAY_HEADERS.get(item, item) for item in headers]]
    type_index = headers.index("row_type") if "row_type" in headers else -1
    side_index = headers.index("side") if "side" in headers else -1
    action_index = headers.index("action") if "action" in headers else -1
    status_index = headers.index("status") if "status" in headers else -1
    for row in rows[1:]:
        mapped = list(row)
        if type_index >= 0 and type_index < len(mapped):
            mapped[type_index] = TYPE_LABELS.get(mapped[type_index], mapped[type_index])
        if side_index >= 0 and side_index < len(mapped):
            mapped[side_index] = normalize_side(mapped[side_index], mapped[type_index] if type_index >= 0 else "")
        if action_index >= 0 and action_index < len(mapped):
            mapped[action_index] = normalize_action(mapped[action_index])
        if status_index >= 0 and status_index < len(mapped):
            mapped[status_index] = normalize_status(mapped[status_index])
        output.append(mapped)
    return headers, output


def normalize_side(value: str, row_type_label: str = "") -> str:
    text = str(value or "").strip()
    if row_type_label == "成交" and text in SIDE_LABELS:
        return SIDE_LABELS[text]
    return SIDE_LABELS.get(text, text)


def normalize_action(value: str) -> str:
    text = str(value or "").strip()
    return ACTION_LABELS.get(text, text)


def normalize_status(value: str) -> str:
    text = str(value or "").strip()
    return STATUS_LABELS.get(text, text)


def sync(csv_path: Path, workbook_path: Path) -> None:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        raise ValueError(f"empty ledger: {csv_path}")

    headers, output_rows = display_rows(rows)
    ledger_rows = [dict(zip(headers, row)) for row in rows[1:]]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "交易总表"
    type_index = headers.index("row_type") if "row_type" in headers else -1
    account_equity_columns = {
        headers.index("start_equity") if "start_equity" in headers else -1,
        headers.index("end_equity") if "end_equity" in headers else -1,
        headers.index("daily_pnl") if "daily_pnl" in headers else -1,
        headers.index("daily_return_pct") if "daily_return_pct" in headers else -1,
    }
    account_equity_columns.discard(-1)
    for row_index, row in enumerate(output_rows):
        sheet.append(display_account_row(row, row_index, type_index, account_equity_columns))

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    type_col = headers.index("row_type") + 1 if "row_type" in headers else 0
    for row in sheet.iter_rows(min_row=2):
        label = row[type_col - 1].value if type_col else ""
        fill = PatternFill("solid", fgColor=TYPE_COLORS[label]) if label in TYPE_COLORS else None
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), 32)
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = max(10, width)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    add_professional_sheets(workbook, ledger_rows)
    add_checkpoint_sheet(workbook, csv_path.parent / "workflow-checkpoints.csv")
    add_account_bridge_sheet(workbook, headers, rows)
    workbook.save(workbook_path)


def display_account_row(row: list[str], row_index: int, type_index: int, account_equity_columns: set[int]) -> list[str]:
    if row_index == 0 or type_index < 0 or row[type_index] == "组合汇总":
        return row
    return ["" if col_index in account_equity_columns else value for col_index, value in enumerate(row)]


def add_professional_sheets(workbook: Workbook, rows: list[dict[str, str]]) -> None:
    add_config_sheet(workbook)
    add_trade_orders_sheet(workbook, rows)
    add_positions_sheet(workbook, rows)
    add_daily_summary_sheet(workbook, rows)
    add_signal_log_sheet(workbook, rows)
    add_audit_log_sheet(workbook, rows)


def add_config_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("配置表")
    sheet.append(["config_key", "config_value", "description"])
    for row in CONFIG_ROWS:
        sheet.append(list(row))
    format_professional_sheet(sheet, title_fill="1F4E79")
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 48
    for cell in sheet["B"]:
        cell.alignment = Alignment(horizontal="right" if cell.row > 1 else "center", vertical="top")


def add_trade_orders_sheet(workbook: Workbook, rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet("交易订单表")
    sheet.append(ORDER_HEADERS)
    trade_index = 1
    plans_by_symbol: dict[str, str] = {}
    for row in rows:
        if row.get("row_type") == "TRADE_PLAN" and row.get("symbol"):
            plans_by_symbol[row["symbol"]] = make_record_id("P", row.get("date", ""), trade_index)
        if row.get("row_type") != "TRADE":
            continue
        date = row.get("date", "")
        order_id = make_record_id("T", date, trade_index)
        side = str(row.get("side", "")).upper()
        sheet.append(
            [
                order_id,
                date,
                row.get("symbol", ""),
                row.get("name", ""),
                side,
                infer_order_type(row),
                numeric_value(row.get("quantity")),
                numeric_value(row.get("price")) if side == "BUY" else None,
                numeric_value(row.get("price")),
                row.get("time", ""),
                normalize_order_status(row.get("status", "")),
                row.get("strategy_id", ""),
                extract_score(row.get("reason", "")),
                row.get("reason", ""),
                plans_by_symbol.get(row.get("symbol", ""), ""),
                date_time_label(date, row.get("time", "")),
            ]
        )
        trade_index += 1
    format_professional_sheet(sheet, title_fill="244062")
    add_list_validation(sheet, "E2:E1048576", ["BUY", "SELL"])
    add_list_validation(sheet, "F2:F1048576", ["LIMIT", "MARKET", "STOP", "CONDITIONAL"])
    add_list_validation(sheet, "K2:K1048576", ["PENDING", "FILLED", "PARTIAL", "CANCELLED", "BLOCKED"])
    apply_number_formats(
        sheet,
        {
            "B": "yyyy-mm-dd",
            "G": "#,##0",
            "H": "#,##0.0000",
            "I": "#,##0.0000",
            "M": "0.0",
            "P": "yyyy-mm-dd hh:mm",
        },
    )
    color_status_column(sheet, "K")
    width_columns(sheet, {"A": 16, "B": 12, "C": 12, "D": 14, "E": 10, "F": 14, "G": 12, "H": 12, "I": 12, "J": 10, "K": 12, "L": 22, "M": 12, "N": 42, "O": 16, "P": 18})


def add_positions_sheet(workbook: Workbook, rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet("持仓明细表")
    sheet.append(POSITION_HEADERS)
    for row in rows:
        if row.get("row_type") != "POSITION":
            continue
        sheet.append(
            [
                row.get("date", ""),
                row.get("symbol", ""),
                row.get("name", ""),
                numeric_value(row.get("quantity")),
                numeric_value(row.get("avg_cost")),
                numeric_value(row.get("price")),
                None,
                None,
                None,
                numeric_value(row.get("weight_pct")) / 100 if numeric_value(row.get("weight_pct")) is not None else None,
                None,
                row.get("notes", ""),
            ]
        )
    max_row = max(sheet.max_row, 2)
    for row_idx in range(2, max_row + 1):
        sheet[f"G{row_idx}"] = f"=D{row_idx}*F{row_idx}"
        sheet[f"H{row_idx}"] = f"=D{row_idx}*(F{row_idx}-E{row_idx})"
        sheet[f"I{row_idx}"] = f'=IFERROR(H{row_idx}/K{row_idx},"")'
        sheet[f"K{row_idx}"] = f"=D{row_idx}*E{row_idx}"
    format_professional_sheet(sheet, title_fill="375623")
    apply_number_formats(
        sheet,
        {
            "A": "yyyy-mm-dd",
            "D": "#,##0",
            "E": "#,##0.0000",
            "F": "#,##0.0000",
            "G": "#,##0.00",
            "H": "#,##0.00;[Red]-#,##0.00;-",
            "I": "0.00%;[Red]-0.00%;-",
            "J": "0.00%",
            "K": "#,##0.00",
        },
    )
    add_gain_loss_format(sheet, f"H2:I{max_row}")
    width_columns(sheet, {"A": 12, "B": 12, "C": 14, "D": 10, "E": 12, "F": 12, "G": 14, "H": 14, "I": 14, "J": 12, "K": 14, "L": 38})


def add_daily_summary_sheet(workbook: Workbook, rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet("每日账户汇总")
    sheet.append(SUMMARY_HEADERS)
    summaries = [row for row in rows if row.get("row_type") == "DAY_SUMMARY"]
    summaries.sort(key=lambda row: (row.get("date", ""), row.get("time", "")))
    for row in summaries:
        sheet.append(
            [
                row.get("date", ""),
                numeric_value(row.get("cash")),
                numeric_value(row.get("stock_value")),
                None,
                None,
                None,
                None,
                numeric_value(row.get("max_drawdown_pct")) / 100 if numeric_value(row.get("max_drawdown_pct")) is not None else None,
                None,
                None,
                None,
                row.get("risk_state", ""),
                0,
                row.get("notes", ""),
            ]
        )
    max_row = max(sheet.max_row, 2)
    for row_idx in range(2, max_row + 1):
        sheet[f"D{row_idx}"] = f"=B{row_idx}+C{row_idx}"
        if row_idx == 2:
            sheet[f"E{row_idx}"] = 0
            sheet[f"F{row_idx}"] = 0
        else:
            sheet[f"E{row_idx}"] = f"=D{row_idx}-D{row_idx-1}"
            sheet[f"F{row_idx}"] = f'=IFERROR(E{row_idx}/D{row_idx-1},"")'
        sheet[f"G{row_idx}"] = f'=IFERROR(D{row_idx}/\'配置表\'!$B$2-1,"")'
        sheet[f"I{row_idx}"] = f'=IFERROR(D{row_idx}/MAX($D$2:D{row_idx})-1,"")'
        sheet[f"K{row_idx}"] = f'=IFERROR(C{row_idx}/D{row_idx},"")'
        sheet[f"L{row_idx}"] = f'=IF(I{row_idx}<=-\'配置表\'!$B$3,"STOP",IF(I{row_idx}<=-\'配置表\'!$B$4,"WARNING","OK"))'
    format_professional_sheet(sheet, title_fill="806000")
    apply_number_formats(
        sheet,
        {
            "A": "yyyy-mm-dd",
            "B": "#,##0.00",
            "C": "#,##0.00",
            "D": "#,##0.00",
            "E": "#,##0.00;[Red]-#,##0.00;-",
            "F": "0.00%;[Red]-0.00%;-",
            "G": "0.00%;[Red]-0.00%;-",
            "H": "0.00%;[Red]-0.00%;-",
            "I": "0.00%;[Red]-0.00%;-",
            "J": "0.00%",
            "K": "0.00%",
            "M": "0.00x",
        },
    )
    color_status_column(sheet, "L")
    width_columns(sheet, {"A": 12, "B": 14, "C": 16, "D": 16, "E": 14, "F": 14, "G": 16, "H": 16, "I": 18, "J": 12, "K": 12, "L": 12, "M": 10, "N": 52})


def add_signal_log_sheet(workbook: Workbook, rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet("策略信号日志")
    sheet.append(SIGNAL_HEADERS)
    signal_index = 1
    for row in rows:
        if row.get("row_type") not in {"TRADE_PLAN", "INTRADAY_CHECK", "STRATEGY_FEEDBACK", "BACKTEST"}:
            continue
        date = row.get("date", "")
        signal_type = infer_signal_type(row)
        actual_action = row.get("action", "")
        sheet.append(
            [
                make_record_id("S", date, signal_index),
                date_time_label(date, row.get("time", "")),
                row.get("symbol", ""),
                row.get("strategy_id") or row.get("strategy_tag", ""),
                signal_type,
                extract_score(row.get("reason", "") + " " + row.get("notes", "")),
                extract_threshold(row.get("score_profile", ""), "entry"),
                extract_threshold(row.get("score_profile", ""), "exit"),
                actual_action,
                actual_action,
                normalize_signal_valid(row.get("status", "")),
                row.get("reason", "") or row.get("notes", ""),
            ]
        )
        signal_index += 1
    format_professional_sheet(sheet, title_fill="5B3F8C")
    add_list_validation(sheet, "E2:E1048576", ["ENTRY", "EXIT", "HOLD", "WATCH", "RESEARCH", "AUDIT"])
    add_list_validation(sheet, "K2:K1048576", ["TRUE", "FALSE", "REVIEW"])
    apply_number_formats(sheet, {"B": "yyyy-mm-dd hh:mm", "F": "0.0", "G": "0.0", "H": "0.0"})
    width_columns(sheet, {"A": 16, "B": 18, "C": 12, "D": 22, "E": 12, "F": 10, "G": 14, "H": 14, "I": 18, "J": 18, "K": 10, "L": 52})


def add_audit_log_sheet(workbook: Workbook, rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet("审计与修正记录")
    sheet.append(AUDIT_HEADERS)
    audit_index = 1
    for row in rows:
        if row.get("row_type") != "AUDIT_FINDING":
            continue
        detail = parse_json_detail(row.get("notes", ""))
        affected_date = str(detail.get("date") or row.get("date") or "")
        symbol = str(detail.get("symbol") or row.get("symbol") or "")
        issue_type = str(detail.get("type") or row.get("reason") or "")
        quantity = detail.get("quantity") or row.get("quantity") or ""
        remaining = detail.get("remaining") or ""
        sheet.append(
            [
                make_record_id("A", row.get("date", ""), audit_index),
                row.get("date", ""),
                issue_type,
                describe_audit_issue(issue_type, symbol, quantity, remaining),
                make_affected_order_id(affected_date, symbol),
                affected_date,
                normalize_audit_status(row.get("status", "")),
                suggested_fix(issue_type),
                "",
                "system",
                row.get("notes", ""),
            ]
        )
        audit_index += 1
    format_professional_sheet(sheet, title_fill="7F1D1D")
    add_list_validation(sheet, "G2:G1048576", ["OPEN", "IN_PROGRESS", "FIXED", "IGNORED", "CLOSED"])
    color_status_column(sheet, "G")
    apply_number_formats(sheet, {"B": "yyyy-mm-dd", "F": "yyyy-mm-dd", "I": "yyyy-mm-dd"})
    width_columns(sheet, {"A": 16, "B": 12, "C": 20, "D": 46, "E": 18, "F": 12, "G": 14, "H": 42, "I": 12, "J": 14, "K": 58})


def make_record_id(prefix: str, date: str, index: int) -> str:
    compact_date = str(date or "").replace("-", "")
    return f"{prefix}{compact_date}{index:03d}" if compact_date else f"{prefix}{index:03d}"


def date_time_label(date: str, time: str) -> str:
    if date and time:
        return f"{date} {time}"
    return date or ""


def numeric_value(value: object) -> float | int | None:
    text = str(value if value is not None else "").strip()
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def infer_order_type(row: dict[str, str]) -> str:
    action = str(row.get("action", "")).upper()
    if "STOP" in action or "止损" in str(row.get("reason", "")):
        return "STOP"
    if row.get("price"):
        return "LIMIT"
    return "MARKET"


def normalize_order_status(value: str) -> str:
    text = str(value or "").strip().lower()
    if text in {"filled", "已成交"}:
        return "FILLED"
    if text in {"partial", "部分成交"}:
        return "PARTIAL"
    if text in {"cancelled", "canceled", "已取消"}:
        return "CANCELLED"
    if text in {"failed", "blocked", "失败"}:
        return "BLOCKED"
    return "PENDING" if text in {"", "pending", "待处理"} else text.upper()


def normalize_signal_valid(value: str) -> str:
    text = str(value or "").strip().lower()
    if text in {"filled", "completed", "done", "已成交", "已完成"}:
        return "TRUE"
    if text in {"failed", "失败"}:
        return "FALSE"
    return "REVIEW"


def normalize_audit_status(value: str) -> str:
    text = str(value or "").strip().lower()
    if text in {"fixed", "closed", "done", "completed", "已完成"}:
        return "CLOSED"
    if text in {"ignored", "waived", "已豁免"}:
        return "IGNORED"
    if text in {"in_progress", "处理中"}:
        return "IN_PROGRESS"
    return "OPEN"


def extract_score(text: str) -> float | None:
    source = str(text or "")
    for marker in ("评分", "score", "Score"):
        if marker not in source:
            continue
        tail = source.split(marker, 1)[1]
        token = ""
        for char in tail:
            if char.isdigit() or char == ".":
                token += char
            elif token:
                break
        if token:
            try:
                return round(float(token), 1)
            except ValueError:
                return None
    return None


def extract_threshold(text: str, key: str) -> float | None:
    source = str(text or "")
    marker = f"{key}="
    if marker not in source:
        return None
    tail = source.split(marker, 1)[1]
    token = ""
    for char in tail:
        if char.isdigit() or char == ".":
            token += char
        elif token:
            break
    if not token:
        return None
    try:
        return round(float(token), 1)
    except ValueError:
        return None


def infer_signal_type(row: dict[str, str]) -> str:
    side = str(row.get("side", "")).upper()
    action = str(row.get("action", "")).upper()
    row_type = row.get("row_type", "")
    if side == "BUY" or "BUY" in action or "买" in row.get("action", ""):
        return "ENTRY"
    if side == "SELL" or "SELL" in action or "卖" in row.get("action", "") or "减仓" in row.get("reason", ""):
        return "EXIT"
    if row_type == "BACKTEST":
        return "RESEARCH"
    if row_type == "TRADE_PLAN":
        return "WATCH"
    return "HOLD"


def parse_json_detail(text: str) -> dict[str, object]:
    source = str(text or "").strip()
    if not source:
        return {}
    try:
        parsed = json.loads(source)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def describe_audit_issue(issue_type: str, symbol: str, quantity: object, remaining: object) -> str:
    if issue_type == "invalid_sell_lot":
        return f"{symbol} 卖出 {quantity} 股非整手且未清仓"
    if issue_type == "odd_lot_remainder":
        return f"{symbol} 卖出后剩余 {remaining} 股零股"
    if issue_type == "t_plus_one_violation":
        return f"{symbol} T+1 可卖数量不足"
    if issue_type == "invalid_buy_lot":
        return f"{symbol} 买入数量不是 100 股整数倍"
    return f"{symbol} {issue_type}".strip()


def suggested_fix(issue_type: str) -> str:
    if issue_type in {"invalid_sell_lot", "odd_lot_remainder"}:
        return "后续执行不得制造零股；不足一手时仅允许清仓或继续持有。"
    if issue_type == "t_plus_one_violation":
        return "按 T+1 可卖数量重算执行包。"
    if issue_type == "invalid_buy_lot":
        return "买入数量向下取整到 100 股整数倍。"
    return "补充人工复核结论。"


def make_affected_order_id(date: str, symbol: str) -> str:
    compact_date = str(date or "").replace("-", "")
    return f"T{compact_date}-{symbol}" if compact_date and symbol else ""


def format_professional_sheet(sheet, title_fill: str = "1F4E79") -> None:
    header_fill = PatternFill("solid", fgColor=title_fill)
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    bottom = Border(bottom=thin_gray)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bottom
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False


def add_list_validation(sheet, cell_range: str, values: list[str]) -> None:
    quoted = ",".join(values)
    validation = DataValidation(type="list", formula1=f'"{quoted}"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def apply_number_formats(sheet, formats: dict[str, str]) -> None:
    for column, number_format in formats.items():
        for cell in sheet[column]:
            if cell.row == 1:
                continue
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)


def width_columns(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def add_gain_loss_format(sheet, cell_range: str) -> None:
    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="F4CCCC")),
    )
    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="D9EAD3")),
    )


def color_status_column(sheet, column: str) -> None:
    cell_range = f"{column}2:{column}{max(sheet.max_row, 2)}"
    sheet.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'OR(${column}2="FILLED",${column}2="OK",${column}2="CLOSED")'], fill=PatternFill("solid", fgColor="D9EAD3")),
    )
    sheet.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'OR(${column}2="WARNING",${column}2="PENDING",${column}2="IN_PROGRESS")'], fill=PatternFill("solid", fgColor="FFF2CC")),
    )
    sheet.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'OR(${column}2="STOP",${column}2="BLOCKED",${column}2="OPEN",${column}2="FAILED")'], fill=PatternFill("solid", fgColor="F4CCCC")),
    )


def add_checkpoint_sheet(workbook: Workbook, checkpoint_path: Path) -> None:
    if not checkpoint_path.exists() or checkpoint_path.stat().st_size == 0:
        return
    rows = list(csv.DictReader(checkpoint_path.open("r", encoding="utf-8-sig", newline="")))
    if not rows:
        return
    sheet = workbook.create_sheet("每日流程检查")
    fields = list(CHECKPOINT_HEADERS)
    sheet.append([CHECKPOINT_HEADERS[field] for field in fields])
    for row in rows:
        sheet.append([row.get(field, "") for field in fields])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_col = fields.index("status") + 1
    for row in sheet.iter_rows(min_row=2):
        status = str(row[status_col - 1].value or "")
        fill = PatternFill("solid", fgColor=STATUS_COLORS[status]) if status in STATUS_COLORS else None
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), 48)
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = max(10, width)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def add_account_bridge_sheet(workbook: Workbook, headers: list[str], rows: list[list[str]]) -> None:
    if not rows:
        return
    sheet = workbook.create_sheet("账户口径说明")
    ledger_rows = [dict(zip(headers, row)) for row in rows[1:]]
    summaries = [row for row in ledger_rows if row.get("row_type") == "DAY_SUMMARY"]
    backtests = [row for row in ledger_rows if row.get("row_type") == "BACKTEST"]
    summaries.sort(key=lambda row: (row.get("date", ""), row.get("time", "")))

    sheet.append(["项目", "日期", "数值", "说明"])
    sheet.append(["正式纸面账户期末权益", "", "", "来自交易总表的组合汇总行；这是账户净值口径。"])
    for row in summaries:
        sheet.append([
            "正式纸面账户期末权益",
            row.get("date", ""),
            row.get("end_equity", ""),
            f"期初={row.get('start_equity', '')}；当日盈亏={row.get('daily_pnl', '')}；现金={row.get('cash', '')}；市值={row.get('stock_value', '')}",
        ])
    for previous, current in zip(summaries, summaries[1:]):
        previous_end = previous.get("end_equity", "")
        current_start = current.get("start_equity", "")
        sheet.append([
            "权益连续性检查",
            current.get("date", ""),
            "通过" if previous_end == current_start else "不一致",
            f"{previous.get('date', '')}期末={previous_end}；{current.get('date', '')}期初={current_start}",
        ])
    for row in backtests:
        sheet.append([
            "回测结果，不计入账户净值",
            row.get("date", ""),
            extract_backtest_metric(row.get("notes", "")),
            "回测是历史策略模拟收益，不是当天纸面账户收益；不得与组合汇总的总资产相加或比较为当日盈亏。",
        ])
    sheet.append([
        "本次疑问解释",
        "",
        "",
        "6/23备注里的约8.7万收益来自回测期末权益1086989.62相对1000000初始资金，不是正式纸面账户当天收益；6/24总资产969598.01来自真实纸面持仓按当日价格结算。",
    ])

    format_simple_sheet(sheet)


def extract_backtest_metric(notes: str) -> str:
    text = str(notes or "")
    parts = []
    for token in text.replace("；", ";").split(";"):
        token = token.strip()
        if any(key in token for key in ("回测期末权益", "总收益", "最大回撤", "Sharpe")):
            parts.append(token)
    return "；".join(parts)


def format_simple_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), 64)
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = max(12, width)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: sync_portfolio_workbook.py <portfolio_ledger.csv> <workbook.xlsx>", file=sys.stderr)
        return 2
    sync(Path(sys.argv[1]), Path(sys.argv[2]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
