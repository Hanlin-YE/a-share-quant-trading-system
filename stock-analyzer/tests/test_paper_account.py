import importlib.util
import pathlib
import sys
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook


SCRIPTS_DIR = pathlib.Path(__file__).resolve().parents[1] / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

MODULE_PATH = SCRIPTS_DIR / "paper_account.py"
SPEC = importlib.util.spec_from_file_location("paper_account", MODULE_PATH)
paper_account = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
sys.modules[SPEC.name] = paper_account
SPEC.loader.exec_module(paper_account)


class PaperAccountTest(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = pathlib.Path(self.tmp.name)
        self.ledger = self.root / "ledger"
        self.account_path = self.root / "account.json"
        self.account_path.write_text(
            """
{
  "initial_cash": 1000000,
  "fee_bps": 3,
  "min_commission": 5,
  "tax_bps": 5,
  "transfer_bps": 0.1,
  "slippage_bps": 0,
  "risk_limits": {
    "max_position_pct": 20,
    "max_total_exposure_pct": 80,
    "daily_loss_stop_pct": 1,
    "max_drawdown_alert_pct": 5,
    "max_drawdown_stop_pct": 10
  }
}
""".strip(),
            encoding="utf-8",
        )
        self.patches = [
            patch.object(paper_account, "JOURNAL_DIR", self.root),
            patch.object(paper_account, "LEDGER_DIR", self.ledger),
            patch.object(paper_account, "ACCOUNT_PATH", self.account_path),
            patch.object(paper_account, "PORTFOLIO_LEDGER_PATH", self.root / "portfolio_ledger.csv"),
            patch.object(paper_account, "WORKBOOK_PATH", self.root / "book.xlsx"),
            patch.object(paper_account, "ORDERS_PATH", self.ledger / "orders.csv"),
            patch.object(paper_account, "POSITIONS_PATH", self.ledger / "positions.csv"),
            patch.object(paper_account, "EQUITY_PATH", self.ledger / "equity_curve.csv"),
        ]
        for item in self.patches:
            item.start()

    def tearDown(self):
        for item in reversed(self.patches):
            item.stop()
        self.tmp.cleanup()

    def test_trade_records_cost_breakdown(self):
        paper_account.ensure_ledger_files()
        row = paper_account.record_trade(
            symbol="600000",
            name="浦发银行",
            side="BUY",
            quantity=100,
            price=9.23,
            trade_date="2026-06-05",
            trade_time="13:00",
            reason="test",
            strategy_tag="stable",
            notes="",
        )

        self.assertEqual(row["commission"], 5)
        self.assertEqual(row["stamp_tax"], 0)
        self.assertGreater(row["total_cost"], 5)
        self.assertLess(row["net_amount"], 0)

    def test_trade_syncs_single_workbook(self):
        paper_account.ensure_ledger_files()
        paper_account.record_trade(
            symbol="600000",
            name="浦发银行",
            side="BUY",
            quantity=100,
            price=9.23,
            trade_date="2026-06-05",
            trade_time="13:00",
            reason="test",
            strategy_tag="stable",
            notes="",
        )

        self.assertTrue(paper_account.WORKBOOK_PATH.exists())
        ws = load_workbook(paper_account.WORKBOOK_PATH)["交易总表"]
        self.assertEqual(ws["A2"].value, "2026-06-05")
        self.assertEqual(ws["C2"].value, "成交")
        self.assertEqual(ws["I2"].value, "100")

    def test_settle_updates_positions_and_equity(self):
        paper_account.ensure_ledger_files()
        paper_account.record_trade(
            symbol="600000",
            name="浦发银行",
            side="BUY",
            quantity=1000,
            price=9.23,
            trade_date="2026-06-05",
            trade_time="13:00",
            reason="test",
            strategy_tag="stable",
            notes="",
        )
        equity = paper_account.settle_day(
            settle_date="2026-06-05",
            prices={"600000": 9.34},
            notes="close",
        )
        positions = paper_account.read_rows(paper_account.POSITIONS_PATH)

        self.assertEqual(len(positions), 1)
        self.assertEqual(positions[0]["symbol"], "600000")
        self.assertGreater(float(equity["end_equity"]), 1_000_000)
        self.assertEqual(equity["risk_state"], "OK")

    def test_single_portfolio_ledger_contains_trade_position_and_day_summary(self):
        paper_account.ensure_ledger_files()
        paper_account.record_trade(
            symbol="600000",
            name="浦发银行",
            side="BUY",
            quantity=1000,
            price=9.23,
            trade_date="2026-06-05",
            trade_time="13:00",
            reason="test",
            strategy_tag="stable",
            notes="",
        )
        paper_account.settle_day(
            settle_date="2026-06-05",
            prices={"600000": 9.34},
            notes="close",
        )

        ledger_rows = paper_account.read_rows(paper_account.PORTFOLIO_LEDGER_PATH)
        row_types = [row["row_type"] for row in ledger_rows]

        self.assertIn("TRADE", row_types)
        self.assertIn("POSITION", row_types)
        self.assertIn("DAY_SUMMARY", row_types)
        self.assertEqual(len([row for row in ledger_rows if row["row_type"] == "DAY_SUMMARY"]), 1)
        self.assertEqual(paper_account.active_positions()["600000"]["quantity"], "1000")
        self.assertGreater(float(paper_account.latest_equity_row()["end_equity"]), 1_000_000)

    def test_t_plus_one_blocks_same_day_sell(self):
        paper_account.ensure_ledger_files()
        paper_account.record_trade(
            symbol="600000",
            name="浦发银行",
            side="BUY",
            quantity=1000,
            price=9.23,
            trade_date="2026-06-05",
            trade_time="13:00",
            reason="test",
            strategy_tag="stable",
            notes="",
        )

        with self.assertRaisesRegex(ValueError, "T\\+1 sell violation"):
            paper_account.record_trade(
                symbol="600000",
                name="浦发银行",
                side="SELL",
                quantity=100,
                price=9.34,
                trade_date="2026-06-05",
                trade_time="14:00",
                reason="same day sell",
                strategy_tag="stable",
                notes="",
            )

    def test_t_plus_one_allows_next_day_sell(self):
        paper_account.ensure_ledger_files()
        paper_account.record_trade(
            symbol="600000",
            name="浦发银行",
            side="BUY",
            quantity=1000,
            price=9.23,
            trade_date="2026-06-05",
            trade_time="13:00",
            reason="test",
            strategy_tag="stable",
            notes="",
        )
        row = paper_account.record_trade(
            symbol="600000",
            name="浦发银行",
            side="SELL",
            quantity=100,
            price=9.34,
            trade_date="2026-06-08",
            trade_time="09:35",
            reason="next day sell",
            strategy_tag="stable",
            notes="",
        )

        self.assertEqual(row["side"], "SELL")

    def test_buy_must_be_board_lot(self):
        paper_account.ensure_ledger_files()

        with self.assertRaisesRegex(ValueError, "BUY quantity must be a multiple"):
            paper_account.record_trade(
                symbol="600000",
                name="浦发银行",
                side="BUY",
                quantity=30,
                price=9.23,
                trade_date="2026-06-05",
                trade_time="13:00",
                reason="invalid buy",
                strategy_tag="stable",
                notes="",
            )

    def test_sell_cannot_create_odd_lot_remainder(self):
        paper_account.ensure_ledger_files()
        paper_account.record_trade(
            symbol="600000",
            name="浦发银行",
            side="BUY",
            quantity=100,
            price=9.23,
            trade_date="2026-06-05",
            trade_time="13:00",
            reason="test",
            strategy_tag="stable",
            notes="",
        )

        with self.assertRaisesRegex(ValueError, "SELL quantity must be a multiple|odd-lot remainder"):
            paper_account.record_trade(
                symbol="600000",
                name="浦发银行",
                side="SELL",
                quantity=30,
                price=9.34,
                trade_date="2026-06-08",
                trade_time="09:35",
                reason="invalid odd lot",
                strategy_tag="stable",
                notes="",
            )

    def test_existing_odd_lot_can_be_closed_in_full(self):
        paper_account.ensure_ledger_files()
        paper_account.append_portfolio_row(
            {
                "date": "2026-06-05",
                "section": "opening",
                "row_type": "OPENING_POSITION",
                "symbol": "600000",
                "name": "浦发银行",
                "side": "BUY",
                "action": "期初持仓同步",
                "quantity": 70,
                "price": 9.23,
                "gross_amount": 646.1,
                "status": "filled",
            }
        )
        paper_account.append_portfolio_row(
            {
                "date": "2026-06-05",
                "section": "summary",
                "row_type": "DAY_SUMMARY",
                "cash": 999000,
                "stock_value": 646.1,
                "end_equity": 999646.1,
            }
        )
        paper_account.append_portfolio_row(
            {
                "date": "2026-06-05",
                "section": "position",
                "row_type": "POSITION",
                "symbol": "600000",
                "name": "浦发银行",
                "side": "LONG",
                "quantity": 70,
                "avg_cost": 9.23,
                "price": 9.23,
                "market_value": 646.1,
                "status": "open",
            }
        )
        row = paper_account.record_trade(
            symbol="600000",
            name="浦发银行",
            side="SELL",
            quantity=70,
            price=9.34,
            trade_date="2026-06-08",
            trade_time="09:35",
            reason="close odd lot",
            strategy_tag="stable",
            notes="",
        )

        self.assertEqual(row["quantity"], 70)

    def test_risk_state_stops_on_daily_loss(self):
        cfg = paper_account.load_account_config()

        self.assertEqual(paper_account.risk_state_for(-1.1, -1.1, cfg), "STOP")
        self.assertEqual(paper_account.risk_state_for(-0.2, -6.0, cfg), "ALERT")
        self.assertEqual(paper_account.risk_state_for(0.1, -1.0, cfg), "OK")


if __name__ == "__main__":
    unittest.main()
