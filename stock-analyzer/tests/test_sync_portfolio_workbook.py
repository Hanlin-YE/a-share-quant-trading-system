import csv
import importlib.util
import pathlib
import sys
import tempfile
import unittest

from openpyxl import load_workbook


ROOT = pathlib.Path(__file__).resolve().parents[2]
TOOLS_DIR = ROOT / "tools"
MODULE_PATH = TOOLS_DIR / "sync_portfolio_workbook.py"
SPEC = importlib.util.spec_from_file_location("sync_portfolio_workbook", MODULE_PATH)
sync_portfolio_workbook = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
sys.modules[SPEC.name] = sync_portfolio_workbook
SPEC.loader.exec_module(sync_portfolio_workbook)


FIELDS = [
    "date",
    "section",
    "row_type",
    "time",
    "symbol",
    "name",
    "side",
    "action",
    "quantity",
    "price",
    "avg_cost",
    "gross_amount",
    "commission",
    "stamp_tax",
    "transfer_fee",
    "total_cost",
    "net_amount",
    "cash",
    "stock_value",
    "market_value",
    "start_equity",
    "end_equity",
    "daily_pnl",
    "daily_return_pct",
    "realized_pnl",
    "unrealized_pnl",
    "unrealized_pnl_pct",
    "total_exposure_pct",
    "weight_pct",
    "max_drawdown_pct",
    "risk_state",
    "strategy_tag",
    "strategy_id",
    "score_profile",
    "factor_weights",
    "strategy_status",
    "reason",
    "thesis",
    "review_date",
    "status",
    "notes",
]


class SyncPortfolioWorkbookTest(unittest.TestCase):
    def write_ledger(self, path, rows):
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=FIELDS)
            writer.writeheader()
            for row in rows:
                writer.writerow({field: row.get(field, "") for field in FIELDS})

    def test_trade_side_is_chinese_and_backtest_does_not_pollute_account_equity(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = pathlib.Path(tmpdir)
            ledger = root / "portfolio_ledger.csv"
            workbook = root / "book.xlsx"
            self.write_ledger(
                ledger,
                [
                    {
                        "date": "2026-06-23",
                        "section": "2026-06-23 交易",
                        "row_type": "TRADE",
                        "symbol": "600000",
                        "side": "BUY",
                        "action": "买入",
                        "status": "filled",
                    },
                    {
                        "date": "2026-06-23",
                        "section": "2026-06-23 组合汇总",
                        "row_type": "DAY_SUMMARY",
                        "start_equity": "975434.47",
                        "end_equity": "975434.47",
                    },
                    {
                        "date": "2026-06-23",
                        "section": "2026-06-23 每日回测",
                        "row_type": "BACKTEST",
                        "side": "RESEARCH",
                        "action": "每日组合回测",
                        "start_equity": "1000000.0",
                        "end_equity": "1086989.62",
                        "status": "completed",
                        "notes": "回测期末权益1086989.62；总收益8.6990%",
                    },
                    {
                        "date": "2026-06-24",
                        "section": "2026-06-24 组合汇总",
                        "row_type": "DAY_SUMMARY",
                        "start_equity": "975434.47",
                        "end_equity": "969598.01",
                    },
                ],
            )

            sync_portfolio_workbook.sync(ledger, workbook)
            wb = load_workbook(workbook)
            ws = wb["交易总表"]

            self.assertEqual(ws["G2"].value, "买入")
            self.assertEqual(ws["AN2"].value, "已成交")
            self.assertEqual(ws["C4"].value, "回测")
            self.assertEqual(ws["AN4"].value, "已完成")
            self.assertIsNone(ws["U4"].value)
            self.assertIsNone(ws["V4"].value)
            self.assertEqual(ws["U3"].value, "975434.47")
            self.assertEqual(ws["V3"].value, "975434.47")
            bridge = wb["账户口径说明"]
            bridge_text = "\n".join(str(cell.value or "") for row in bridge.iter_rows() for cell in row)
            self.assertIn("回测结果，不计入账户净值", bridge_text)
            self.assertIn("权益连续性检查", bridge_text)

    def test_pending_validation_status_is_rendered_as_overdue_evidence_gap(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = pathlib.Path(tmpdir)
            ledger = root / "portfolio_ledger.csv"
            workbook = root / "book.xlsx"
            self.write_ledger(
                ledger,
                [
                    {
                        "date": "2026-06-05",
                        "section": "2026-06-05 13:30复查",
                        "row_type": "INTRADAY_CHECK",
                        "review_date": "待14:00验证",
                        "status": "待验证",
                    },
                ],
            )

            sync_portfolio_workbook.sync(ledger, workbook)
            ws = load_workbook(workbook)["交易总表"]

            self.assertEqual(ws["AN2"].value, "逾期未留证")

    def test_audit_finding_stock_name_column_stays_empty(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = pathlib.Path(tmpdir)
            ledger = root / "portfolio_ledger.csv"
            workbook = root / "book.xlsx"
            self.write_ledger(
                ledger,
                [
                    {
                        "date": "2026-06-26",
                        "section": "2026-06-26 账本审计",
                        "row_type": "AUDIT_FINDING",
                        "symbol": "600519",
                        "name": "",
                        "side": "RESEARCH",
                        "action": "审计发现",
                        "quantity": "30",
                    },
                ],
            )

            sync_portfolio_workbook.sync(ledger, workbook)
            ws = load_workbook(workbook)["交易总表"]

            self.assertEqual(ws["C2"].value, "审计发现")
            self.assertEqual(ws["E2"].value, "600519")
            self.assertIsNone(ws["F2"].value)
            self.assertEqual(ws["H2"].value, "审计发现")

    def test_professional_views_are_derived_from_single_ledger(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = pathlib.Path(tmpdir)
            ledger = root / "portfolio_ledger.csv"
            workbook = root / "book.xlsx"
            self.write_ledger(
                ledger,
                [
                    {
                        "date": "2026-06-25",
                        "section": "2026-06-25 盘后策略计划",
                        "row_type": "TRADE_PLAN",
                        "time": "15:40",
                        "symbol": "002230",
                        "name": "科大讯飞",
                        "side": "SELL",
                        "action": "明日减仓计划",
                        "quantity": "100",
                        "price": "41.50",
                        "strategy_id": "score-standard-v1",
                        "score_profile": "entry=58 exit=45",
                        "reason": "评分 38.5 低于退出阈值45",
                        "status": "pending",
                    },
                    {
                        "date": "2026-06-26",
                        "section": "2026-06-26 交易",
                        "row_type": "TRADE",
                        "time": "09:30",
                        "symbol": "002230",
                        "name": "科大讯飞",
                        "side": "SELL",
                        "action": "卖出",
                        "quantity": "100",
                        "price": "40.3498",
                        "strategy_id": "score-standard-v1",
                        "reason": "评分 38.5 低于退出阈值 45.0",
                        "status": "filled",
                    },
                    {
                        "date": "2026-06-27",
                        "section": "2026-06-27 持仓",
                        "row_type": "POSITION",
                        "symbol": "002230",
                        "name": "科大讯飞",
                        "side": "LONG",
                        "quantity": "200",
                        "price": "40.37",
                        "avg_cost": "42.7184",
                        "weight_pct": "0.834",
                    },
                    {
                        "date": "2026-06-27",
                        "section": "2026-06-27 组合汇总",
                        "row_type": "DAY_SUMMARY",
                        "cash": "841254.12",
                        "stock_value": "126814.5",
                        "end_equity": "968068.62",
                        "risk_state": "OK",
                    },
                    {
                        "date": "2026-06-29",
                        "section": "2026-06-29 账本审计",
                        "row_type": "AUDIT_FINDING",
                        "symbol": "300750",
                        "side": "RESEARCH",
                        "action": "审计发现",
                        "quantity": "30",
                        "reason": "invalid_sell_lot",
                        "status": "failed",
                        "notes": '{"type":"invalid_sell_lot","date":"2026-06-24","symbol":"300750","quantity":30,"remaining":70}',
                    },
                ],
            )

            sync_portfolio_workbook.sync(ledger, workbook)
            wb = load_workbook(workbook, data_only=False)

            for sheet_name in [
                "交易订单表",
                "持仓明细表",
                "每日账户汇总",
                "策略信号日志",
                "审计与修正记录",
                "配置表",
            ]:
                self.assertIn(sheet_name, wb.sheetnames)

            orders = wb["交易订单表"]
            self.assertEqual(orders["A2"].value, "T20260626001")
            self.assertEqual(orders["E2"].value, "SELL")
            self.assertEqual(orders["K2"].value, "FILLED")

            positions = wb["持仓明细表"]
            self.assertEqual(positions["G2"].value, "=D2*F2")
            self.assertEqual(positions["I2"].value, '=IFERROR(H2/K2,"")')

            summary = wb["每日账户汇总"]
            self.assertEqual(summary["D2"].value, "=B2+C2")
            self.assertEqual(summary["L2"].value, '=IF(I2<=-\'配置表\'!$B$3,"STOP",IF(I2<=-\'配置表\'!$B$4,"WARNING","OK"))')

            audit = wb["审计与修正记录"]
            self.assertEqual(audit["C2"].value, "invalid_sell_lot")
            self.assertEqual(audit["G2"].value, "OPEN")


if __name__ == "__main__":
    unittest.main()
